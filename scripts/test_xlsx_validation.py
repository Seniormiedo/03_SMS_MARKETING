#!/usr/bin/env python3
"""
Test script para verificar que los números de validación aparezcan en archivos XLSX
"""

import sys
import os
import asyncio
from pathlib import Path
import openpyxl

# Add bot directory to Python path
sys.path.insert(0, str(Path(__file__).parent.parent / 'bot'))

from services.contact_service import ContactService
from services.export_service import ExportService
from models.extraction import ExtractionRequest, ExtractionType, ExportFormat

async def test_xlsx_validation():
    """Test validation numbers in XLSX export"""
    print("🔍 PROBANDO NÚMEROS DE VALIDACIÓN EN ARCHIVOS XLSX")
    print("=" * 60)
    
    try:
        # Initialize services
        print("1. Inicializando servicios...")
        contact_service = ContactService()
        export_service = ExportService()
        
        await contact_service.initialize()
        await export_service.initialize()
        print("✅ Servicios inicializados")
        
        # Create extraction request for 100 contacts first (no validation numbers)
        print("\n2. Creando solicitud de extracción (100 contactos)...")
        request = ExtractionRequest(
            extraction_type=ExtractionType.PREMIUM,
            amount=100,
            location=None,
            export_format=ExportFormat.XLSX
        )
        print(f"✅ Solicitud creada: {request.amount} contactos, formato {request.export_format}")
        
        # Extract contacts
        print("\n3. Extrayendo contactos con números de validación...")
        result = await contact_service.extract_contacts(request)
        
        if result.status != "COMPLETED":
            print(f"❌ Error en extracción: {result.error_message}")
            return False
            
        print(f"✅ Extracción exitosa:")
        print(f"  - Total extraídos: {len(result.contacts)}")
        print(f"  - Contactos reales: {result.total_found}")
        print(f"  - Números de validación: {len(result.contacts) - result.total_found}")
        
        # Check for validation contacts
        validation_contacts = [c for c in result.contacts if hasattr(c, 'state_name') and c.state_name == 'VALIDACION']
        print(f"  - Contactos de validación detectados: {len(validation_contacts)}")
        
        if validation_contacts:
            print("  - Números de validación:")
            for vc in validation_contacts:
                print(f"    * {vc.phone_national} → {vc.get_display_location()}")
        
        # Export to XLSX
        print("\n4. Exportando a archivo XLSX...")
        file_path = await export_service.export_contacts(result)
        print(f"✅ Archivo XLSX creado: {file_path}")
        
        # Verify XLSX content
        print("\n5. Verificando contenido del archivo XLSX...")
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"✅ Archivo XLSX abierto, filas: {worksheet.max_row}")
        
        # Look for validation numbers in the file
        validation_rows = []
        for row in range(2, worksheet.max_row + 1):  # Skip header
            content = worksheet[f'B{row}'].value
            if content and "VALIDACION" in str(content):
                phone = worksheet[f'A{row}'].value
                validation_rows.append((row, phone, content))
        
        print(f"✅ Números de validación encontrados en XLSX: {len(validation_rows)}")
        
        if validation_rows:
            print("  - Números en archivo XLSX:")
            for row_num, phone, content in validation_rows:
                print(f"    Fila {row_num}: {phone} → {content}")
        else:
            print("❌ NO se encontraron números de validación en el archivo XLSX")
            
            # Debug: Show first few rows
            print("\nPrimeras 5 filas del archivo para debug:")
            for row in range(1, min(6, worksheet.max_row + 1)):
                phone = worksheet[f'A{row}'].value
                content = worksheet[f'B{row}'].value
                print(f"  Fila {row}: {phone} → {content}")
        
        # Clean up
        workbook.close()
        
        print("\n" + "=" * 60)
        if validation_rows:
            print("✅ NÚMEROS DE VALIDACIÓN APARECEN CORRECTAMENTE EN XLSX")
            return True
        else:
            print("❌ NÚMEROS DE VALIDACIÓN NO APARECEN EN XLSX")
            return False
            
    except Exception as e:
        print(f"\n❌ ERROR EN PRUEBA: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = asyncio.run(test_xlsx_validation())
    sys.exit(0 if success else 1)
