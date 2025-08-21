"""
Export service for Contact Extractor Bot
Handles file generation in XLSX and TXT formats
"""

import time
from datetime import datetime
from pathlib import Path
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from config import get_config
from utils.logger import get_logger
from utils.formatters import PhoneFormatter, FileNameGenerator, DataFormatter
from models.contact import Contact
from models.extraction import ExtractionResult, ExportFormat


class ExportService:
    """
    Professional service for exporting contacts to files
    Supports XLSX and TXT formats with proper formatting
    """
    
    def __init__(self):
        self.config = get_config()
        self.logger = get_logger()
        self._initialized = False
    
    async def initialize(self):
        """Initialize the export service"""
        if self._initialized:
            return
        
        try:
            # Ensure export directory exists
            export_path = Path(self.config.export_path)
            export_path.mkdir(parents=True, exist_ok=True)
            
            self._initialized = True
            self.logger.info("✅ Export service initialized")
            
        except Exception as e:
            self.logger.error(f"Failed to initialize export service: {e}")
            raise
    
    async def export_contacts(self, result: ExtractionResult) -> str:
        """
        Export contacts to file based on request format
        
        Args:
            result: Extraction result with contacts and request info
            
        Returns:
            str: Path to generated file
        """
        if not result.contacts:
            raise ValueError("No contacts to export")
        
        start_time = time.time()
        
        try:
            # Generate filename
            filename = FileNameGenerator.generate_extraction_filename(
                str(result.request.extraction_type),
                len(result.contacts),
                result.request.location,
                str(result.request.export_format)
            )
            
            file_path = Path(self.config.export_path) / filename
            
            # Export based on format
            if result.request.export_format == ExportFormat.XLSX:
                await self._export_to_xlsx(result.contacts, file_path)
            elif result.request.export_format == ExportFormat.TXT:
                await self._export_to_txt(result.contacts, file_path)
            else:
                raise ValueError(f"Unsupported export format: {result.request.export_format}")
            
            export_time = time.time() - start_time
            result.export_time_seconds = export_time
            
            # Get file size
            file_size = file_path.stat().st_size
            
            # Complete the result
            result.mark_completed(str(file_path), file_size)
            
            self.logger.info(
                f"Export completed: {filename} ({file_size:,} bytes) in {export_time:.2f}s"
            )
            
            # Log file export for audit
            self.logger.log_file_export(
                str(file_path),
                file_size,
                str(result.request.export_format)
            )
            
            return str(file_path)
            
        except Exception as e:
            error_msg = f"Export failed: {str(e)}"
            self.logger.error(error_msg)
            result.mark_failed(error_msg)
            raise
    
    async def _export_to_xlsx(self, contacts: List[Contact], file_path: Path):
        """
        Export contacts to Excel file
        
        Args:
            contacts: List of contacts to export
            file_path: Output file path
        """
        try:
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = self.config.xlsx_sheet_name
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            data_alignment = Alignment(horizontal="left", vertical="center")
            
            # Add headers
            worksheet['A1'] = "Number"
            worksheet['B1'] = "Content"
            
            # Style headers
            for cell in ['A1', 'B1']:
                worksheet[cell].font = header_font
                worksheet[cell].fill = header_fill
                worksheet[cell].alignment = header_alignment
            
            # Add data
            for row, contact in enumerate(contacts, start=2):
                # Format phone to 12 digits
                phone_12 = PhoneFormatter.format_to_digits(contact.phone_national, 12)
                
                # Get location (city or fallback)
                location = contact.get_display_location()
                
                # Truncate Content column to maximum 11 characters
                content_truncated = location[:11] if location else ""
                
                # Log truncation if content was cut
                if location and len(location) > 11:
                    self.logger.info(f"Content truncated: '{location}' -> '{content_truncated}'")
                
                # Add to worksheet
                worksheet[f'A{row}'] = phone_12
                worksheet[f'B{row}'] = content_truncated
                
                # Style data cells
                worksheet[f'A{row}'].alignment = data_alignment
                worksheet[f'B{row}'].alignment = data_alignment
            
            # Auto-adjust column widths
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 25
            
            # Add metadata sheet
            meta_sheet = workbook.create_sheet("Metadata")
            meta_sheet['A1'] = "Generated"
            meta_sheet['B1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            meta_sheet['A2'] = "Total Contacts"
            meta_sheet['B2'] = len(contacts)
            meta_sheet['A3'] = "Format"
            meta_sheet['B3'] = "12-digit phone numbers"
            meta_sheet['A4'] = "Content Format"
            meta_sheet['B4'] = "Truncated to 11 characters max"
            meta_sheet['A5'] = "Bot Version"
            meta_sheet['B5'] = self.config.bot_version
            
            # Save workbook
            workbook.save(file_path)
            
            self.logger.debug(f"XLSX file created: {file_path}")
            
        except Exception as e:
            self.logger.error(f"XLSX export failed: {e}")
            raise
    
    async def _export_to_txt(self, contacts: List[Contact], file_path: Path):
        """
        Export contacts to text file
        
        Args:
            contacts: List of contacts to export
            file_path: Output file path
        """
        try:
            with open(file_path, 'w', encoding=self.config.txt_encoding) as file:
                for contact in contacts:
                    # Format phone to 12 digits
                    phone_12 = PhoneFormatter.format_to_digits(contact.phone_national, 12)
                    file.write(f"{phone_12}\n")
            
            self.logger.debug(f"TXT file created: {file_path}")
            
        except Exception as e:
            self.logger.error(f"TXT export failed: {e}")
            raise
    
    def cleanup_old_files(self, days_old: int = None):
        """
        Clean up old export files
        
        Args:
            days_old: Files older than this many days will be deleted
        """
        if days_old is None:
            days_old = self.config.file_retention_days
        
        try:
            export_path = Path(self.config.export_path)
            if not export_path.exists():
                return
            
            cutoff_time = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
            deleted_count = 0
            
            for file_path in export_path.iterdir():
                if file_path.is_file() and file_path.stat().st_mtime < cutoff_time:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                    except Exception as e:
                        self.logger.warning(f"Failed to delete {file_path}: {e}")
            
            if deleted_count > 0:
                self.logger.info(f"Cleaned up {deleted_count} old export files")
            
        except Exception as e:
            self.logger.error(f"Cleanup failed: {e}")
    
    def get_export_stats(self) -> dict:
        """
        Get export statistics
        
        Returns:
            dict: Export statistics
        """
        try:
            export_path = Path(self.config.export_path)
            if not export_path.exists():
                return {"total_files": 0, "total_size_mb": 0}
            
            files = list(export_path.glob("*"))
            total_size = sum(f.stat().st_size for f in files if f.is_file())
            
            return {
                "total_files": len(files),
                "total_size_mb": round(total_size / (1024 * 1024), 2),
                "export_path": str(export_path),
                "retention_days": self.config.file_retention_days
            }
            
        except Exception as e:
            self.logger.error(f"Failed to get export stats: {e}")
            return {"error": str(e)}


# Export main class
__all__ = ["ExportService"]